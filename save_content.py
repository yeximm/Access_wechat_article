import requests
import os
import re       # 使用正则表达式
import time
import json     # 用于json转码
import random
import openpyxl
from bs4 import BeautifulSoup


class SaveContent():
    """
        完成网页验证
        获取单个文章的网页文本数据
        保存单个文章的网页为pdf格式(待实现)
    """
    def __init__(self):
        self.save_root_path = r'./data'
        self.names = '公众号默认名称（需更换）'      # 程序默认以‘公众号名称’做目录，用以保存相关公众号的信息
        self.headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) '
                          'Chrome/116.0.0.0 Safari/537.36 NetType/WIFI MicroMessenger/7.0.20.1781(0x6700143B) '
                          'WindowsWechat(0x63090a13) XWEB/9129 Flue',
        }
        self.cookies = {"poc_sid": ''}     # 用以保存设备ID
        os.makedirs(self.save_root_path, exist_ok=True)  # 创建保存路径，如果文件夹已存在，则忽略，默认为r'./data'

    def get_all_content(self, content):
        """
            汇总所有的信息
            :param content:网页内容
        """
        title = self.get_title(content)
        times = self.get_time(content)
        texts = self.get_texts(content)
        return {
            'title': title,     # 文章标题
            'times': times,     # 文章发布时间
            'texts': texts,     # 文章内容（列表形式）
            'all_type': '',  # 文章包含的标志信息，用于获取文章阅读量等信息
        }

    def verify_user(self, url, content):
        """
            url=请求路径
            content=网页内容，如：res.text
            遇到此情况时使用：  >当前环境异常，完成验证后即可继续访问。<
            返回参数；验证标志（1为有效），网页内容，cookie值
                {'verify_flag': 1, 'content': res.text, 'poc_sid': poc_sid}
        """
        print('开始验证，正在获取参数poc_sid')
        poc_token = re.search(r'poc_token.*"(.*?)"', content).group(1)
        poc_sid = re.search(r'poc_sid.*"(.*?)"', content).group(1)  # poc_sid为cookie参数
        cap_appid = re.search(r'cap_appid.*"(.*?)"', content).group(1)
        cap_sid = re.search(r'cap_sid.*"(.*?)"', content).group(1)
        target_url = re.search(r'target_url.*"(.*?)"', content).group(1)

        try:
            '''验证请求第一步'''
            verify1_url = ('https://t.captcha.qq.com/cap_union_prehandle?'
                           'aid=' + cap_appid +
                           '&protocol=https&accver=1&showtype=popup&ua=TW96aWxsYS81LjAgKFdpbmRvd3MgTlQgMTAuMDsgV2luNjQ7IHg2NCkgQXBwbGVXZWJLaXQvNTM3LjM2IChLSFRNTCwgbGlrZSBHZWNrbykgQ2hyb21lLzEyNy4wLjAuMCBTYWZhcmkvNTM3LjM2IEVkZy8xMjcuMC4wLjA%3D&noheader=0&fb=1&aged=0&enableAged=0&enableDarkMode=1&'
                           'deviceID=' + poc_sid + '&sid=' + cap_sid +
                           '&grayscale=1&dyeid=0&clientype=2&cap_cd=&uid=&lang=zh-cn&entry_url=https%3A%2F%2Fmp.weixin.qq.com%2Fmp%2Fwappoc_appmsgcaptcha&elder_captcha=0&js=%2Ftcaptcha-frame.8d77d8b0.js&login_appid=&wb=1&version=1.1.0&subsid=1&callback=_aq_873604&sess=')
            verify1 = requests.get(verify1_url, headers=self.headers, verify=False)
            # sess，pow_answer 为第三步中的请求数据
            sess = re.search('sess":"(.*?)"', verify1.text).group(1)
            pow_answer = re.search('prefix":"(.*?)"', verify1.text).group(1)

            '''验证请求第二步'''
            verify2_url = 'https://t.captcha.qq.com' + re.search('tdc_path":"(.*?)"', verify1.text).group(1)
            verify2 = requests.get(verify2_url, headers=self.headers, verify=False)
            # eks 为第三步中的请求数据
            eks = re.search(r"='(.*?)'", verify2.text).group(1)

            '''验证请求第三步'''
            verify3_url = 'https://t.captcha.qq.com/cap_union_new_verify'
            verify3_data = {
                # 'collect': 'rTdL2csh9MFPSaeKi0kIljgBgCZKq9QgnKBxrlpTn3YT41x6hx3sIjZVFn544i8YgxUt4iyQSxZpkD%2BeFzrQB4MzlnMqochzx0%2BrFwXMMEnmz5ec2RysL1qlKV9BAQz5K3gw7jWpBPJWtyWKD5UWTK39jnbMtitCUGRFdQQS0XYeIX3GYdrmSJc0a4JdPx3YuZW5YTdar%2BhbG07vWc8IsCWv9WYfmFkfvXzUWm5alHEuoKngABAFUVZWlvaulCpZ89U5cI90c9Cuz4gVc5q3bXOTNk1hWo3dF3CfnbVnoezWefsjIzyR9Fn1fNK2nv7rVMYUdPORoo5WVpb2rpQqWVZWlvaulCpZC5095cBnxOUbkZU9ko3C0CF3qrZy0qXgg%2BcsfLgi6mn63I2eOkQD%2FISjcH3WeTtN37dGnsQOeLGofkN2MCNpMirzT2sFN4ZOImNS%2FY6FQOGZ7UQsyRg8xly%2FO5ChYYsBMIru78AYuJi2tACDtOxCi0Sepo7US8w8LwtCIyNOUQQO4531ay7Hrexsw4KgeneWb2cj%2FLzXypA2Vga97ex9S61bW%2B1BAHTpDglMOA5VC5VbrVAau6mh7HkWP90YufkLAfKedGQRWpTyXiDJlKhbu5OC3ZAcJXDtHvI0uQtGD6j4y0edGMc0bT9zDNW33WLsL%2BEVpVv7AuHl9VlmZ0Zc5Q8BO0yDe70FQgIpI5F%2BWoKhUxQDqAtbL2WeCGqbj9AG9PAZkMr620aAsIGBSCufUG22gnWrAFUJ6okHj5Q7kYycBYBoYjkFkv8WAof1Z5WkOkwQ%2FUGx7NyOY5VyKZPqEcj5n2r7r4rL2Ha35M23XH73CDno6e6QHeSXsYdUw2%2BG4l7Rp8QNyAlzzVeAP69v7oZ%2BjuAxxmLBwEtFRk0qVnlW%2F%2BZYj7LybHtpL%2FSt9V2egQiO1Y5qZ1z36gbbnborLA2%2BEQIQYd01SZnsl7OsO0cViJ4wBEcfgCOMBDvnJBfyenM1nRY3zaeuecnhbJIzK33ROtYsD3Y7o%2FKYghPdictG32I59qyo5qn2gG4aPaWoTOOBa9jQU3nJy6TRgtpw1p1zsthUoo%2B3Z9YxZlYEJZXdsG2797jLExc7vg3BwXxKfxIUezrobzkABVVH%2FTYkNnraXUv8GAtvAmFdyPK7y%2FBYzBUw9SfmYGAvemDFbjc0kesd2Pl7rME51FURZdu8AW%2Bni6reapdxg4u5REEfUaZbXT7bE4aqwIMPqcezvluHIghv3dCZZ0n%2Fn1zNADEjB1ZWlvaulCpZcg0Mp9p9ZhTR9MfLjjO6to6e2sUK9OCUkI6fIkvjG0ooDtXbCdfD5Q%3D%3D',
                # 'tlg': 1336,
                'eks': eks,
                'sess': sess,
                'ans': '[{"elem_id":0,"type":"DynAnswerType_TIME","data":""}]',
                'deviceID': poc_sid,
                'pow_answer': pow_answer,
                # 'pow_calc_time': 3,
            }
            verify3 = requests.post(verify3_url, headers=self.headers, data=verify3_data, verify=False)
            # ticket，randstr为第四步请求参数
            ticket = json.loads(verify3.text)['ticket']
            randstr = json.loads(verify3.text)["randstr"]

            '''验证请求第四步'''
            verify4_url = 'https://mp.weixin.qq.com/mp/wappoc_appmsgcaptcha?action=Check&x5=0&f=json'
            verify4_data = {
                'target_url': target_url,
                'poc_token': poc_token,
                'appid': cap_appid,
                'ticket': ticket,
                'randstr': randstr,
            }
            self.cookies['poc_sid'] = poc_sid   # 重置类属性 cooikes的值
            print('获取到poc_sid：'+ self.cookies['poc_sid'])
            verify4 = requests.post(verify4_url, headers=self.headers, cookies=self.cookies, data=verify4_data, verify=False)
            # print(verify4.text)
            # print('发送成功后，poc_sid就可以正常使用了')

            '''验证请求第五步'''
            modify_url = url + '&poc_token=' + poc_token
            res = requests.get(modify_url, headers=self.headers, cookies=self.cookies, verify=False)  # 发起请求
            print('已完成验证请求，后续请求可正常使用。')
            return {'verify_flag': 1, 'content': res.text, 'poc_sid': poc_sid}
        except:
            print('验证失败，请检查后再进行尝试')
            return {'verify_flag': 0}

    def get_content(self, url):
        """
            输入微信文章链接（永久链接或临时链接）
            返回参数：请求标志位（1时有效），文章内容（网页内容），请求参数cookie
            {'content_flag': 1, 'content': res.text, 'poc_sid': self.cookies['poc_sid']}
        """
        res = requests.get(url, headers=self.headers, cookies=self.cookies, verify=False)  # 发起请求
        # 验证请求
        if 'var createTime = ' in res.text:     # 正常获取到文章内容
            print('正常获取到文章内容，当前cookie值为：poc_sid=' + self.cookies['poc_sid'])
            # self.names = re.search(r'var nickname.*"(.*?)".*', res.text).group(1)  # 公众号名称
            return {'content_flag': 1, 'content': res.text, 'poc_sid': self.cookies['poc_sid']}
        elif '>当前环境异常，完成验证后即可继续访问。<' in res.text:
            print('当前环境异常，完成验证后即可继续访问。')
            # 返回环境异常，需进行验证操作
            contents = self.verify_user(url, res.text)
            if contents['verify_flag'] == 1:
                # self.names = re.search(r'var nickname.*"(.*?)".*', contents['content']).group(1)  # 公众号名称
                return {'content_flag': 1, 'content': contents['content'], 'poc_sid': contents['poc_sid']}
            else:
                print('验证操作失败！！！！！请检查后重试')
                return {'content_flag': 0}
        elif '操作频繁，请稍后再试。' in res.text:
            # 锁cookie了，等待几个小时后再尝试
            print('操作频繁了，等会再弄或换ip弄！！！')
            print(res.text)
            return {'content_flag': 0}
        else:   # 出现错误信息，待处理！！！！！！！！！
            print("请求出错，请等待几秒后请继续操作")
            print('出现其他问题，请查找原因后再试！！！！')
            return {'content_flag': 0}

    def get_img(self, content, title, creat_time):
        """
            将文章中的图片下载，并存储到本地
            存储路径在 self.save_root_path 下
            搭配create_directory_and_download_image()使用
        """
        imgs = content.split('https://mmbiz.qpic.cn/')
        # 在Windows操作系统中，文件名需要遵守以下命名规则：
        # 文件名不包含以下任何字符：“（双引号）、*（星号）、<（小于）、>（大于）、？（问号）、\（反斜杠）、/（正斜杠）、|（竖线）、：（冒号）。
        # 文件名不要以空格、句号、连字符或下划线开头或结尾。
        title = title.replace('"', '“').replace('*', '※').replace('|', 'I').replace('/', ' ')
        title = title.replace('\\', ' ').replace(':', '：').replace('<', '《').replace('>', '》').replace('?', '？')
        img_save_path = self.names + "/" + creat_time + '_' + title    # 图片保存路径

        for i in range(0, len(imgs) - 1):
            img_url = 'https://mmbiz.qpic.cn/' + imgs[i + 1].split('"')[0]
            # print('正在获取图片：' + img_url)
            self.create_directory_and_download_image(img_url, self.save_root_path + "/" + img_save_path, str(i + 1))

    def create_directory_and_download_image(self, image_url, directory_path,img_name):
        image_name = ''
        os.makedirs(directory_path, exist_ok=True)
        response = requests.get(image_url, cookies=self.cookies, verify=False)
        if response.status_code == 200:
            img_hz = ['gif', 'jpg']
            for imghz in img_hz:
                if imghz in image_url:
                    image_name = img_name + '.' + imghz
            if image_name == '':
                image_name = img_name + '.jpg'
            file_path = directory_path + '/' + image_name
            with open(file_path, 'wb') as f:
                f.write(response.content)
            print(f"已成功下载图片： {file_path}")
        else:
            print(f"无法下载图片，状态码: {response.status_code}")

    def get_title(self, content):
        """
            文章标题在<h1></h1>标签内
            '.' 匹配任意字符，除了换行符，当re.DOTALL标记被指定时，则可以匹配包括换行符的任意字符。
            re.search(path, content, re.DOTALL).group(1)获取的内容为
                class="rich_media_title " id="activity-name">title
            split('>')将获取到的内容从'>'开始切片
            strip()删除字符串两边空白
        """
        path = r'<h1(.*)</h1'
        tittle = re.search(path, content, re.DOTALL).group(1).split('>')[1].strip()      # strip()去除字符串中空字符
        return tittle

    def get_time(self, content):
        """
            尝试获取时间，程序到此处时已经过验证，可以得到时间信息
            timex为正则表达式筛选出的数据：2024-08-01 07:09
            返回参数（例）： ['2024-08-01', '07:09']
        """
        timex = re.search(r'var createTime = (.*?) .*', content).group().split("'")[1]
        year = str(timex.split(' ')[0].split('-')[0])
        month = str(timex.split(' ')[0].split('-')[1])
        day = str(timex.split(' ')[0].split('-')[2])
        hour = str(timex.split(' ')[1])
        # detail_time = str(year) + '年' + str(month) + '月' + str(day) + '日'
        detail_time = [timex.split(' ')[0], hour]
        return detail_time

    def get_texts(self, content):
        """
            将文字内容转换为列表形式存储
        """
        soup = BeautifulSoup(content, 'html.parser')
        spans = soup.find_all('p')
        texts = [span.text for span in spans]
        # print('正在读取页面内容：' + str(texts))
        return texts

    def read_excel_content(self, data_path):
        # 打开Excel文件
        workbook = openpyxl.load_workbook(data_path)
        # 获取活动工作表（通常是第一个工作表）
        sheet = workbook.active
        # 读取整行或整列数据
        # row_values = [cell.value for cell in sheet[1]]  # 第一行数据
        column_id = [sheet[f'A{i}'].value for i in range(2, sheet.max_row + 1)]  # A列数据，数据id
        column_values = [sheet[f'C{i}'].value for i in range(2, sheet.max_row + 1)]  # C列数据，文章标题
        return column_id, column_values

    def creat_excel_content(self, save_path, pxlsx):
        # 保存 微信文章内容.xlsx
        exist_excel = True  # 不存在为True
        if os.path.exists(save_path + pxlsx):   # 检测文件是否存在，若不存在则新创建
            print(save_path + pxlsx + "，该文件存在")
            column_id, title_list = self.read_excel_content(save_path + pxlsx)
            if title_list:  # 表中有数据
                print('表中存在数据，即将更新文件内容')
                exist_excel = False
            else:   # 表中无数据，重置表格内容
                print('注意！表中无数据，将重置表格内容')
        else:
            print(save_path + pxlsx + '文件不存在，将创建新文件')
        if exist_excel:
            wb = openpyxl.Workbook()
            wb.save(save_path + pxlsx)
            array = ['序号', '发布时间', '文章名称', '文章链接', '文章内容']
            array_1 = ['A', 'B', 'C', 'D', 'E']
            for i in range(0, len(array)):
                w = openpyxl.load_workbook(save_path + pxlsx)
                sheet = w.active
                sheet[array_1[i] + "1"] = array[i]
                w.save(save_path + pxlsx)

    def write_excel_content(self, path, orders, times, titles, links, texts):
        wb = openpyxl.load_workbook(path)
        sheet = wb.active
        sheet['A' + str(orders + 1)] = orders
        sheet['B' + str(orders + 1)] = times
        sheet['C' + str(orders + 1)] = titles
        sheet['D' + str(orders + 1)] = links
        sheet['E' + str(orders + 1)] = texts
        wb.save(path)
        order = orders + 1
        return order


class GetList(SaveContent):
    """
        ① 请求得到文章信息（文章标题、文章链接、文章创建日期）
        ②以excel文件形式存储，文件名设置为对应公众号的名称
    """
    def __init__(self):
        super().__init__()
        self.biz = 'MzkxMDMxODEwOA=='   # 默认为国务院客户端公众号，后续可重新赋值
        self.uin = ''
        self.key = ''
        self.pass_ticket = ''
        self.appmsg_token = ''

    def save_main(self, pagess):
        passage_list = self.get_passage_list(pagess)
        if passage_list['message_flag'] == 1:   # 文章信息获取成功
            # print(passage_list)
            if passage_list['passage_list'][0][0]:    # 存在至少一篇文章
                content_url = passage_list['passage_list'][0][2]
                # 调用save_content文件，获取公众号名称
                sac = SaveContent()
                sac.get_content(content_url)
                os.makedirs(self.save_root_path + '/' + sac.names, exist_ok=True)  # 创建data文件夹，如果文件夹已存在，则忽略
                self.save_list(self.save_root_path + '/' + sac.names, passage_list)    # 保存文章列表
                return {'name_flag': 1, 'names': sac.names}
            else:
                print('该公众号下没有文章，，，')
                return {'name_flag': 0}
        else:
            # print(passage_list)
            print('文章信息获取失败')
            return {'name_flag': 0}

    def get_passage_list(self, pagess):
        """
            将每页的文章信息汇总
        """
        # 获取文章链接，文章标题，文章创建时间
        passage_list = []
        message_flag = 1
        for pages in range(pagess):
            # print('pages=' + str(pages))
            p_data = self.get_message_new(pages)
            # return titles_links['all_titles']
            if p_data['m_flag'] == 1:
                for i in p_data['passage_list']:
                    passage_list.append(i)
            else:
                print('请求出错！页面解析出现问题')
                message_flag = 0
                break
            delay_time = random.randint(1, 3)   # 延迟时间
            print('为预防被封禁,开始延时操作，延时时间：' + str(delay_time) + '秒')
            time.sleep(delay_time)    # 模拟手动操作，随机延时1-3秒，预防被封禁
        if message_flag == 1:
            return {
                'message_flag': 1,
                'length': len(passage_list),
                'passage_list': passage_list,
            }
        else:   # 出错了
            return {'message_flag': 0}

    def get_list_all(self):
        # 获取公众号下所有的文章链接
        page = 0
        passage_list = []
        print('开始获取公众号下所有的文章列表')
        while True:
            p_data = self.get_message_new(page)
            # return titles_links['all_titles']
            if p_data['m_flag'] == 1:
                for i in p_data['passage_list']:
                    passage_list.append(i)
            else:
                print('请求结束，文章列表获取完毕！')
                break
            page = page + 1
            delay_time = random.randint(1, 3)  # 延迟时间
            print('为预防被封禁,开始延时操作，延时时间：' + str(delay_time) + '秒')
            time.sleep(delay_time)  # 模拟手动操作，随机延时1-3秒，预防被封禁
        if passage_list:
            return {
                'message_flag': 1,
                'length': len(passage_list),
                'passage_list': passage_list,
            }
        else:  # 出错了
            return {'message_flag': 0}

    def get_message_new(self,page):
        # 从0开始计数，第 0页相当于默认页数据
        pages = int(page) * 10
        print('正在获取第 ' + str(page+1) + ' 页文章列表')
        url = ('https://mp.weixin.qq.com/mp/profile_ext?action=getmsg&__biz=' + self.biz + '&f=json&offset='
               + str(pages) + '&count=10&is_ok=1&scene=124&uin=' + self.uin + '&key=' + self.key + '&pass_ticket='
               + self.pass_ticket + '&wxtoken=&appmsg_token=' + self.appmsg_token + '&x5=0&f=json')
        res = requests.get(url=url, headers=self.headers, timeout=10, verify=False)
        # print(res.text)
        if 'app_msg_ext_info' in res.text:
            # 解码json数据
            get_page = json.loads(json.loads(res.text)['general_msg_list'])['list']
            ''' get_page[0]为
            {'comm_msg_info': {'id': 1000000107, 'type': 49, 'datetime': 1722467332, 'fakeid': '3910318108', 'status': 2, 'content': ''}, 'app_msg_ext_info': {'title': '国务院7月重要政策', 'digest': '', 'content': '', 'fileid': 100007840, 'content_url': 'http://mp.weixin.qq.com/s?__biz=MzkxMDMxODEwOA==&amp;mid=2247491511&amp;idx=1&amp;sn=a36291fdee52a0f53d145edec8058e04&amp;chksm=c0084d6abbcac962a50153c89fe9c19b6f8b1c5e5ac50b05adcb49bdfad8638522ab426c3f4b&amp;scene=27#wechat_redirect', 'source_url': '', 'cover': 'https://mmbiz.qpic.cn/mmbiz_jpg/JRAjbHqmggrlZibDMibLP4ryNqhYXgolJOdQj2P8t2QQFVicickzAo7Gv1SzazwJY6lDylcanx2ic60HDbMvK8OKQpg/0?wx_fmt=jpeg', 'subtype': 9, 'is_multi': 1, 'multi_app_msg_item_list': [{'title': '8月起，这些新规将影响你我生活！', 'digest': '', 'content': '', 'fileid': 0, 'content_url': 'http://mp.weixin.qq.com/s?__biz=MzkxMDMxODEwOA==&amp;mid=2247491511&amp;idx=2&amp;sn=b3f5b6bcf8727c8c90fce7e588e6e7da&amp;chksm=c0eb20c99ca2f90032a6234002ed2cc9c2c000f87cff34f4d8d763878c0bb5275800db876ca7&amp;scene=27#wechat_redirect', 'source_url': '', 'cover': 'https://mmbiz.qpic.cn/mmbiz_jpg/JRAjbHqmggrc08yJMZ6CQ3VL6VzmEIymSUyATlL6o3xaDJJ0D2CtpQg31Vy7jdCaic86zqkgJ9oAFGyia78ZOq7g/0?wx_fmt=jpeg', 'author': '', 'copyright_stat': 100, 'del_flag': 1, 'item_show_type': 0, 'audio_fileid': 0, 'duration': 0, 'play_url': '', 'malicious_title_reason_id': 0, 'malicious_content_type': 0}, {'title': '8月，你好！', 'digest': '', 'content': '', 'fileid': 100007860, 'content_url': 'http://mp.weixin.qq.com/s?__biz=MzkxMDMxODEwOA==&amp;mid=2247491511&amp;idx=3&amp;sn=cd25de57b74b63b0f3b1a9888b9cd94d&amp;chksm=c0c7f30fdd5fc0ea4a2765f5fd29e1faeb0e352e888ee8556521ab23bc9528d68f42deaa9d15&amp;scene=27#wechat_redirect', 'source_url': '', 'cover': 'https://mmbiz.qpic.cn/mmbiz_jpg/JRAjbHqmggrlZibDMibLP4ryNqhYXgolJO9CnECAnMLDPY39Y9iarcFtM1ibrBvhKcGFyl1wicHysvTrYx4GfLybt8g/0?wx_fmt=jpeg', 'author': '', 'copyright_stat': 100, 'del_flag': 1, 'item_show_type': 0, 'audio_fileid': 0, 'duration': 0, 'play_url': '', 'malicious_title_reason_id': 0, 'malicious_content_type': 0}], 'author': '', 'copyright_stat': 100, 'duration': 0, 'del_flag': 1, 'item_show_type': 0, 'audio_fileid': 0, 'play_url': '', 'malicious_title_reason_id': 0, 'malicious_content_type': 0}}
            存储形式为二维数组，[[时间，文章标题，文章链接],[时间，文章标题，文章链接]
            '''
            passage_list = []   # 存放一页内的所有文章
            for i in get_page:
                # 时间戳转换
                time_tuple = time.localtime(i['comm_msg_info']['datetime'])
                create_time = time.strftime("%Y-%m-%d", time_tuple)
                title = i['app_msg_ext_info']['title']
                content_url = i['app_msg_ext_info']['content_url'].replace('#wechat_redirect', '')
                passage_list.append([create_time, title, content_url])
                if i['app_msg_ext_info']['multi_app_msg_item_list']:
                    for j in i['app_msg_ext_info']['multi_app_msg_item_list']:
                        title = j['title']
                        content_url = j['content_url'].replace('#wechat_redirect', '')
                        passage_list.append([create_time, title, content_url])
            # print(passage_list, len(passage_list))
            return {
                'm_flag': 1,
                'passage_list': passage_list,
                'length': len(passage_list)
            }
        elif '"home_page_list":[]' in res.text:
            print('\n出现：操作频繁，请稍后再试\n该号已被封禁，请解封后再来！！！\n')
            return {'m_flag': 0}
        else:
            print('请求结束！未获取到第 ' + str(page + 1) + ' 页文章列表')
            return {'m_flag': 0}

    def save_list(self, data_path, data):
        """
            将文章列表保存到文件
            只存4项数据：序号、时间、文章名称、文章链接
        """
        filename = '/文章列表.xlsx'
        save_path = data_path + filename
        exert = self.creat_excel_url(data_path, filename)   # 创建excel表
        if exert['exist_excel'] == 0:   # 文件不存在时，直接写入信息
            for i in range(data['length']):
                times = data['passage_list'][i][0]
                title = data['passage_list'][i][1]
                content_url = data['passage_list'][i][2]
                self.write_excel_url(save_path, i+1, times, title, content_url)
                print('新写入文章数据：' + title)
        else:   # 此时文件存放在目标处
            # 读取列表内容，返回标题和序号的列表
            column_id, title_list = self.read_excel_url(save_path)
            current_id = len(title_list) + 1    # 用于新添数据时使用
            # 遍历数据，如果已存在，则写入，若不存在，则在末尾添加新数据
            for i in data['passage_list']:
                if i[1] in title_list:  #若有标题已经在表中，更新表中数据
                    times = i[0]
                    title = i[1]
                    content_url = i[2]
                    text_id = int(title_list.index(title)) + 1
                    self.write_excel_url(save_path, text_id, times, title, content_url)
                    print('已更新文章数据：' + title)
                else:   # 此时标题不在列表中
                    # 此时current_id处为空行
                    times = i[0]
                    title = i[1]
                    content_url = i[2]
                    self.write_excel_url(save_path, current_id, times, title, content_url)
                    print('在序号为：' + str(current_id) + ' 处添加新文章：' + title)
                    current_id += 1
                    print('添加完后，序号变为：' + str(current_id))
        print('\n文章链接等信息已保存在：' + save_path)

    def read_excel_url(self, data_path):
        # 打开Excel文件
        workbook = openpyxl.load_workbook(data_path)
        # 获取活动工作表（通常是第一个工作表）
        sheet = workbook.active
        # 读取整行或整列数据
        # row_values = [cell.value for cell in sheet[1]]  # 第一行数据
        column_id = [sheet[f'A{i}'].value for i in range(2, sheet.max_row + 1)]  # A列数据
        column_values = [sheet[f'C{i}'].value for i in range(2, sheet.max_row + 1)]  # C列数据
        return column_id, column_values

    def creat_excel_url(self, save_path, pxlsx):
        if os.path.exists(save_path + pxlsx):   # 检测文件是否存在，若不存在则新创建
            print("列表文件存在")
            column_id, title_list = self.read_excel_url(save_path + pxlsx)
            if title_list:  # 表中有数据
                return {'exist_excel': 1}
            else:   # 表中无数据，重置表格内容
                print('注意！表中无数据，将重置表格内容')
                wb = openpyxl.Workbook()
                wb.save(save_path + pxlsx)
                array = ['序号', '发布时间', '文章名称', '文章链接']
                array_1 = ['A', 'B', 'C', 'D']
                for i in range(0, len(array)):
                    w = openpyxl.load_workbook(save_path + pxlsx)
                    sheet = w.active
                    sheet[array_1[i] + "1"] = array[i]
                    w.save(save_path + pxlsx)
                return {'exist_excel': 0}
        else:
            print('文件不存在')
            wb = openpyxl.Workbook()
            wb.save(save_path + pxlsx)
            array = ['序号', '发布时间', '文章名称', '文章链接']
            array_1 = ['A', 'B', 'C', 'D']
            for i in range(0, len(array)):
                w = openpyxl.load_workbook(save_path + pxlsx)
                sheet = w.active
                sheet[array_1[i] + "1"] = array[i]
                w.save(save_path + pxlsx)
            return {'exist_excel': 0}

    def write_excel_url(self, path, orders, times, titles, links):
        wb = openpyxl.load_workbook(path)
        sheet = wb.active
        sheet['A' + str(orders + 1)] = orders
        sheet['B' + str(orders + 1)] = times
        sheet['C' + str(orders + 1)] = titles
        sheet['D' + str(orders + 1)] = links
        wb.save(path)
        order = orders + 1
        return order





