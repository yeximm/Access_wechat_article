from get_detail import SaveAllDetail
import openpyxl
import os
import re
import random
import time
import urllib3
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)     # 忽略https证书提示
sad = SaveAllDetail()
"""
        得到公众号主页链接，粘贴到微信开始抓包，使用fiddler工具获取到请求的关键字值
        biz值：是微信公众号的标识码，每个公众号都有一个唯一的biz值
        uin值：就是user information，也就是特定微信用户的ID，通过fiddler抓到的包中获取
        key值：是一个动态参数
        pass_ticket值：微信登录之后返回的参数
"""


def get_article_link(url):
    """
        输入公众号下任意一篇已发布的文章 短链接！！
        通过公众号内的文章获取到公众号的biz值，拼接出公众号主页链接
    """
    content = sad.get_content(url)
    if content['content_flag'] == 1:
        print('正在生成微信公众号主页链接……\n')
        biz = re.search('__biz=(.*?)&', content['content']).group(1)
        names = re.search(r'var nickname.*"(.*?)".*', content['content']).group(1)  # 公众号名称
        main_url = ('https://mp.weixin.qq.com/mp/profile_ext?action=home&__biz=' + biz + '&scene=124#wechat_redirect')
        return {'link_flag': 1, 'main_url': main_url, 'names': names}
    else:
        print('未获取到文章内容，请检查链接是否正确')
        return {'link_flag': 0,}


def save_article_list(biz, uin, key, pass_ticket, pages=0):
    """
        获取微信公众号列表，保存到指定文件夹
        biz, uin, key, pass_ticket为固定参数，以此来通过验证
        pages 为公众号下文章页数，从1开始计数
        在此步创建以 公众号名称 命名的文件夹
    """
    sad.biz = biz
    sad.uin = uin
    sad.key = key
    sad.pass_ticket = pass_ticket

    '''①确定是否为获取全部的文章列表'''
    if pages == 0:
        print('没有输入值，开始获取全部文章列表')
        all_list = sad.get_list_all()
    else:
        print('输入值为：' + str(pages) + '，开始获取前' + str(pages) + '页文章')
        all_list = sad.get_passage_list(pages)

    '''②获取公众号名称，通过第一篇文章链接得到公众号名称'''
    if all_list['message_flag'] == 1:
        print('已获取到文章列表，共' + str(all_list['length']) + '篇文章')
        first_list = all_list['passage_list'][0][2]
        # print('列表中第一个链接为：' + first_list)
        content = sad.get_content(first_list)
        sad.names = re.search(r'var nickname.*"(.*?)".*', content['content']).group(1)  # 公众号名称
        sad.cookies['poc_sid'] = content['poc_sid']     # 写入cookie值到临时类变量中
        print('公众号名称为：' + sad.names)

        '''③创建保存路径，将获取的数据进行保存'''
        os.makedirs(r'./data/' + sad.names, exist_ok=True)  # 创建data文件夹，如果文件夹已存在，则忽略
        print('文件数据保存目录为：' + r'./data/' + sad.names)
        sad.save_list(r'./data/' + sad.names, all_list)  # 保存文章列表
    else:
        print('未获取到数据，请重新获取链接后再试')


def save_article_content(save_path, save_img):
    """
        保存微信公众号列表内的文章正文内容
        save_path：文章保存路径，如：r'./data/国务院客户端/'
        读取文件：文章列表.xlsx
        创建文件：微信文章内容.xlsx        # 该文件表头有5项：序号、发布时间、文章名称、文章链接、文章内容
        保存每篇文章内容到：save_path/<公众号名称>/微信文章内容.xlsx
    """
    # 输入文件地址，读取文章标题、文章链接
    column_title, column_url = [], []
    if os.path.exists(save_path) and os.path.exists(save_path + '文章列表.xlsx'):    # 检查文件及其目录是否存在，若不存在则报错！
        print("“文章列表.xlsx”文件存在，正在读取：" + save_path + '文章列表.xlsx')
        workbook = openpyxl.load_workbook(save_path + '文章列表.xlsx')  # 打开：文章列表.xlsx
        sheet1 = workbook.active  # 获取活动工作表（通常是第一个工作表）
        column_title = [sheet1[f'C{i}'].value for i in range(2, sheet1.max_row + 1)]  # 读取C列数据，文章标题
        column_url = [sheet1[f'D{i}'].value for i in range(2, sheet1.max_row + 1)]  # 读取D列数据，文章链接
        if column_url:
            print('获取到文章列表,一共' + str(len(column_title)) + '篇文章\n开始保存各篇文章')
            sad.creat_excel_content(save_path, '微信文章内容.xlsx')  # 创建 微信文章内容.xlsx，将文章内容进行保存操作
            for i, j in zip(column_url, range(len(column_title))):
                # j为表头id，写入到'/微信文章内容.xlsx'时使用
                content = sad.get_content(i)
                if content['content_flag'] == 1:
                    detail_time = sad.get_time(content['content'])  # 获取文章发布时间信息 ['2024-08-01', '07:09']
                    texts = sad.get_texts(content['content'])  # 列表形式的文章内容
                    tittle = sad.get_title(content['content'])  # 文章标题
                    print('正在下载文章：' + tittle)
                    if save_img:
                        sad.get_img(content['content'], tittle, detail_time[0])  # 自动下载文章图片

                    '''此处获取到文章信息，开始进行保存操作'''
                    sad.write_excel_content(save_path + '微信文章内容.xlsx', j + 1, detail_time[0], tittle, i, str(texts))
                else:
                    print('获取文章内容失败，请检查问题后再继续')
        else:
            print('文章列表.xlsx 为空，请先获取文章的链接信息，保存在‘文章列表.xlsx’内')
    else:
        print('未在本地查询到该公众号的文章信息，请先获取文章的链接信息，保存在‘文章列表.xlsx’内')


def save_article_detail(biz, uin, key, pass_ticket):
    """
        保存微信公众号文章的全部内容
        biz, uin, key, pass_ticket为固定参数，以此来通过验证
        poc_sid 为临时参数，获取文章内容时使用
        save_path 为对应公众号文章的存储路径
    """
    sad.biz = biz
    sad.uin = uin
    sad.key = key
    sad.pass_ticket = pass_ticket
    first_list = sad.get_passage_list(1)        # 查看公众号下第一页文章
    first_url = first_list['passage_list'][0][2]    # 获取到名下第一篇文章链接，报错时检查sad.get_passage_list
    content = sad.get_content(first_url)
    sad.names = re.search(r'var nickname.*"(.*?)".*', content['content']).group(1)  # 公众号名称
    sad.cookies['poc_sid'] = content['poc_sid']  # 写入cookie值到临时类变量中
    print('公众号名称为：' + sad.names)
    print('文件数据保存目录为：' + r'./data/' + sad.names)
    # 输入文件地址，读取文章标题、文章链接
    column_title, column_url = [], []
    save_path = sad.save_root_path + '/' + sad.names + '/'
    if os.path.exists(save_path) and os.path.exists(save_path + '文章列表.xlsx'):  # 检查文件及其目录是否存在，若不存在则报错！
        print("“文章列表.xlsx”文件存在，正在读取：" + save_path + '文章列表.xlsx')
        workbook = openpyxl.load_workbook(save_path + '文章列表.xlsx')  # 打开：文章列表.xlsx
        sheet1 = workbook.active  # 获取活动工作表（通常是第一个工作表）
        column_title = [sheet1[f'C{i}'].value for i in range(2, sheet1.max_row + 1)]  # 读取C列数据，文章标题
        column_url = [sheet1[f'D{i}'].value for i in range(2, sheet1.max_row + 1)]  # 读取D列数据，文章链接
        if column_url:
            print('获取到文章列表,一共' + str(len(column_title)) + '篇文章\n开始保存各篇文章')
            '''开始保存所有文章'''
            sad.creat_excel_detail(save_path, '微信文章全部内容.xlsx')  # 创建文章内容.xlsx，将文章内容进行保存操作
            for i, j, z in zip(column_title, column_url, range(len(column_title))):
                # print(getdetail.get_detail_new(i,j))
                (read_num, like_num, share_num, looking_num, comments,
                 like_numss, detail_time, title, texts) = sad.get_detail_new(j, i)

                '''此处获取到文章信息，开始进行保存操作'''
                print('正在保存文章：' + i)
                sad.write_excel_detail(save_path + '微信文章全部内容.xlsx', z + 1, detail_time[0], title,  read_num,
                                       like_num, share_num, looking_num, j, str(comments),  str(like_numss), str(texts))
                delay_time = random.randint(1, 4)  # 延迟时间，生成一个介于a和b之间的随机整数，包括a和b
                print('为预防被封禁,开始延时操作，延时时间：' + str(delay_time) + '秒\n')
                time.sleep(delay_time)  # 模拟手动操作，随机延时1-3秒，预防被封禁
        else:
            print('文章列表.xlsx 为空，请先获取文章的链接信息，保存在‘文章列表.xlsx’内')
    else:
        print('未在本地查询到该公众号的文章信息，请先获取文章的链接信息，保存在‘文章列表.xlsx’内')


if __name__=="__main__":
    root_path = sad.save_root_path + '/'     # 默认存储路径./data
    screen_text = '''请输入数字键！
        数字键1：获取公众号主页链接（输入公众号下任意一篇已发布的文章链接即可）
        数字键2：获取公众号下文章列表（每页文章有15篇）
        数字键3：下载文章内容，自动下载文章列表中所有文章内容
        数字键4：同功能3，下载文章内容，包括单个文章的文本内容 + 阅读量、点赞数等信息
                （请注意请求间隔，若请求太多太快可能会触发封禁！！）
    输入其他任意字符退出！'''
    print('欢迎使用，' + screen_text)
    while True:
        text = str(input('请输入功能数字：'))

        if text == '1':
            random_url = (input('（默认公众号主页链接为“研招网资讯”，按回车键使用）\n请输入公众号下任意一篇已发布的文章链接：') or
                          'https://mp.weixin.qq.com/s/4r_LKJu0mOeUc70ZZXK9LA')
            main_link = get_article_link(random_url)
            print(main_link['names'] + '主页链接为：' + main_link['main_url'])
            print('将此链接 （￣︶￣）↗　 粘贴发送到 ‘微信PC端-文件传输助手’')
            print('\n' + screen_text)

        elif text == '2':
            texts = input('\n以下内容需要用到fiddler工具！！！！！\n（1）在微信客户端打开步骤1获取到的链接，\n'
                  '（2）在fiddler中查看——主机地址为https://mp.weixin.qq.com，URL地址为：/mp/profile_ext?acti\n'
                  '（3）选中此项后按快捷键：Ctrl+U，复制此网址到剪贴板\n（4）将该内容粘贴到此处 (づ￣ 3￣)づ\n请输入复制的链接：')
            biz = re.search('biz=(.*?)&', texts)
            uin = re.search('uin=(.*?)&', texts)
            key = re.search('key=(.*?)&', texts)
            pass_ticket = re.search('pass_ticket=(.*?)&', texts)
            if biz and uin and pass_ticket and key:
                pages = input('（默认获取全部历史文章）\n请输入需要下载的最新发布文章的页数(例：1)：') or 0
                save_article_list(biz.group(1), uin.group(1), key.group(1), pass_ticket.group(1), int(pages))
            else:
                print('\n※※※ 请输入符合要求的链接！')
            print('\n' + screen_text)

        elif text == '3':
            change_name, save_img = '', ''
            print('默认保存路径：' + root_path)
            text_names3 = input('检测到当前的微信公众号名称为：' + sad.names + '，是否更换公众号？\n' +
                                '是(请输入微信公众号名称，例如：泰山风景名胜区)，否(默认，直接按回车跳过)')
            if text_names3:  # 输入新名称
                sad.names = change_name
                change_name = text_names3
            else:
                print('未更换公众号名称')
                change_name = sad.names
            save_path = root_path + change_name + '/'
            print('当前保存路径为：' + save_path)
            save_img = input('是否保存图片？是(输入任意值)，否(默认，直接按回车跳过)————(y/N)')
            save_article_content(save_path, save_img)
            print('\n' + screen_text)

        elif text == '4':
            texts = input('\n以下内容需要用到fiddler工具！！！！！\n（1）在微信客户端打开步骤1获取到的链接，\n'
                          '（2）在fiddler中查看——主机地址为https://mp.weixin.qq.com，URL地址为：/mp/profile_ext?acti\n'
                          '（3）选中此项后按快捷键：Ctrl+U，复制此网址到剪贴板\n（4）将该内容粘贴到此处 (づ￣ 3￣)づ\n请输入复制的链接：')
            biz = re.search('biz=(.*?)&', texts)
            uin = re.search('uin=(.*?)&', texts)
            key = re.search('key=(.*?)&', texts)
            pass_ticket = re.search('pass_ticket=(.*?)&', texts)
            if biz and uin and pass_ticket and key:
                save_article_detail(biz.group(1), uin.group(1), key.group(1), pass_ticket.group(1))
            else:
                print('\n※※※ 请输入符合要求的链接！')
            print('\n' + screen_text)

        else:
            print('\n已成功退出！')
            break










