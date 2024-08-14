import time
import requests
import json
import os
import openpyxl
from save_content import SaveContent


class GetUser(SaveContent):
    """
        ① 请求得到文章信息（文章标题、文章链接、文章创建日期）
        ②以excel文件形式存储，文件名设置为对应公众号的名称
    """
    def __init__(self):
        super().__init__()
        self.biz = 'MzkxMDMxODEwOA=='   # 默认为国务院客户端公众号
        self.uin = ''
        self.key = ''
        self.pass_ticket = ''
        self.appmsg_token = ''

    def seve_main(self, pagess):
        passage_list = self.get_passage_list(pagess)
        if passage_list['message_flag'] == 1:   # 文章信息获取成功
            # print(passage_list)
            if passage_list['passage_list'][0][0]:    # 存在至少一篇文章
                content_url = passage_list['passage_list'][0][2]
                # 调用save_content文件，获取公众号名称
                sac = SaveContent()
                sac.get_content(content_url)
                os.makedirs(r'./data/' + sac.names, exist_ok=True)  # 创建data文件夹，如果文件夹已存在，则忽略
                self.save_list(r'./data/' + sac.names, passage_list)    # 保存文章列表
                return {'name_flag': 1, 'names': sac.names}
            else:
                print('该公众号下没有文章，，，')
                return {'name_flag': 0}
        else:
            # print(passage_list)
            print('文章信息获取失败')
            return {'name_flag': 0}

    def get_passage_list(self, pagess):
        """
            将每页的文章信息汇总
        """
        # 获取文章链接，文章标题，文章创建时间
        passage_list = []
        message_flag = 1
        for pages in range(1, pagess + 1):
            # print('pages=' + str(pages))
            p_data = self.get_message_new(pages)
            # return titles_links['all_titles']
            if p_data['m_flag'] == 1:
                for i in p_data['passage_list']:
                    passage_list.append(i)
            else:
                print('请求出错！页面解析出现问题')
                message_flag = 0
                break
        if message_flag == 1:
            return {
                'message_flag': 1,
                'length': len(passage_list),
                'passage_list': passage_list,
            }
        else:   # 出错了
            return {'message_flag': 0}

    def get_message_new(self,page):
        get_Alltitle = []
        get_Alllink = []
        pages = int(page - 1) * 10
        url = ('https://mp.weixin.qq.com/mp/profile_ext?action=getmsg&__biz=' + self.biz + '&f=json&offset='
               + str(pages) + '&count=10&is_ok=1&scene=124&uin=' + self.uin + '&key=' + self.key + '&pass_ticket='
               + self.pass_ticket + '&wxtoken=&appmsg_token=' + self.appmsg_token + '&x5=0&f=json')
        res = requests.get(url=url, headers=self.headers, timeout=10, verify=False)
        # print(res.text)
        if 'app_msg_ext_info' in res.text:
            # 解码json数据
            # get_page = jsonpath.jsonpath(json.loads(res.text), "$.." + 'general_msg_list')
            # for i in get_page:
            #     get_title = jsonpath.jsonpath(json.loads(i), "$.." + 'title')
            #     get_link = jsonpath.jsonpath(json.loads(i), "$.." + 'content_url')
            #     for titles in get_title:
            #         get_Alltitle.append(titles)
            #     for links in get_link:
            #         get_Alllink.append(links)

            get_page = json.loads(json.loads(res.text)['general_msg_list'])['list']
            ''' get_page[0]为
            {'comm_msg_info': {'id': 1000000107, 'type': 49, 'datetime': 1722467332, 'fakeid': '3910318108', 'status': 2, 'content': ''}, 'app_msg_ext_info': {'title': '国务院7月重要政策', 'digest': '', 'content': '', 'fileid': 100007840, 'content_url': 'http://mp.weixin.qq.com/s?__biz=MzkxMDMxODEwOA==&amp;mid=2247491511&amp;idx=1&amp;sn=a36291fdee52a0f53d145edec8058e04&amp;chksm=c0084d6abbcac962a50153c89fe9c19b6f8b1c5e5ac50b05adcb49bdfad8638522ab426c3f4b&amp;scene=27#wechat_redirect', 'source_url': '', 'cover': 'https://mmbiz.qpic.cn/mmbiz_jpg/JRAjbHqmggrlZibDMibLP4ryNqhYXgolJOdQj2P8t2QQFVicickzAo7Gv1SzazwJY6lDylcanx2ic60HDbMvK8OKQpg/0?wx_fmt=jpeg', 'subtype': 9, 'is_multi': 1, 'multi_app_msg_item_list': [{'title': '8月起，这些新规将影响你我生活！', 'digest': '', 'content': '', 'fileid': 0, 'content_url': 'http://mp.weixin.qq.com/s?__biz=MzkxMDMxODEwOA==&amp;mid=2247491511&amp;idx=2&amp;sn=b3f5b6bcf8727c8c90fce7e588e6e7da&amp;chksm=c0eb20c99ca2f90032a6234002ed2cc9c2c000f87cff34f4d8d763878c0bb5275800db876ca7&amp;scene=27#wechat_redirect', 'source_url': '', 'cover': 'https://mmbiz.qpic.cn/mmbiz_jpg/JRAjbHqmggrc08yJMZ6CQ3VL6VzmEIymSUyATlL6o3xaDJJ0D2CtpQg31Vy7jdCaic86zqkgJ9oAFGyia78ZOq7g/0?wx_fmt=jpeg', 'author': '', 'copyright_stat': 100, 'del_flag': 1, 'item_show_type': 0, 'audio_fileid': 0, 'duration': 0, 'play_url': '', 'malicious_title_reason_id': 0, 'malicious_content_type': 0}, {'title': '8月，你好！', 'digest': '', 'content': '', 'fileid': 100007860, 'content_url': 'http://mp.weixin.qq.com/s?__biz=MzkxMDMxODEwOA==&amp;mid=2247491511&amp;idx=3&amp;sn=cd25de57b74b63b0f3b1a9888b9cd94d&amp;chksm=c0c7f30fdd5fc0ea4a2765f5fd29e1faeb0e352e888ee8556521ab23bc9528d68f42deaa9d15&amp;scene=27#wechat_redirect', 'source_url': '', 'cover': 'https://mmbiz.qpic.cn/mmbiz_jpg/JRAjbHqmggrlZibDMibLP4ryNqhYXgolJO9CnECAnMLDPY39Y9iarcFtM1ibrBvhKcGFyl1wicHysvTrYx4GfLybt8g/0?wx_fmt=jpeg', 'author': '', 'copyright_stat': 100, 'del_flag': 1, 'item_show_type': 0, 'audio_fileid': 0, 'duration': 0, 'play_url': '', 'malicious_title_reason_id': 0, 'malicious_content_type': 0}], 'author': '', 'copyright_stat': 100, 'duration': 0, 'del_flag': 1, 'item_show_type': 0, 'audio_fileid': 0, 'play_url': '', 'malicious_title_reason_id': 0, 'malicious_content_type': 0}}
            存储形式为二维数组，[[时间，文章标题，文章链接],[时间，文章标题，文章链接]
            '''
            passage_list = []   # 存放一页内的所有文章
            for i in get_page:
                # 时间戳转换
                time_tuple = time.localtime(i['comm_msg_info']['datetime'])
                create_time = time.strftime("%Y-%m-%d", time_tuple)
                title = i['app_msg_ext_info']['title']
                content_url = i['app_msg_ext_info']['content_url'].replace('#wechat_redirect', '')
                passage_list.append([create_time, title, content_url])
                if i['app_msg_ext_info']['multi_app_msg_item_list']:
                    for j in i['app_msg_ext_info']['multi_app_msg_item_list']:
                        title = j['title']
                        content_url = j['content_url'].replace('#wechat_redirect', '')
                        passage_list.append([create_time, title, content_url])
            # print(passage_list, len(passage_list))
            return {
                'm_flag': 1,
                'passage_list': passage_list,
                # 'length': len(passage_list)
            }
        else:
            print("请求失败！未获取到文章列表")
            return {'m_flag': 0}

    def save_list(self, data_path, data):
        """
            将文章列表保存到文件
            只存4项数据：序号、时间、文章名称、文章链接
        """
        filename = '/文章列表.xlsx'
        save_path = data_path + filename
        exert = self.creat_excel(data_path, filename)   # 创建excel表
        if exert['exist_excel'] == 0:   # 文件不存在时，直接写入信息
            for i in range(data['length']):
                times = data['passage_list'][i][0]
                title = data['passage_list'][i][1]
                content_url = data['passage_list'][i][2]
                self.write_excel_detail(save_path, i+1, times, title, content_url)
                print('新写入文章数据：' + title)
        else:   # 此时文件存放在目标处
            # 读取列表内容，返回标题和序号的列表
            column_id, title_list = self.read_excel(save_path)
            current_id = len(title_list) + 1    # 用于新添数据时使用
            # 遍历数据，如果已存在，则写入，若不存在，则在末尾添加新数据
            for i in data['passage_list']:
                if i[1] in title_list:  #若有标题已经在表中，更新表中数据
                    times = i[0]
                    title = i[1]
                    content_url = i[2]
                    text_id = int(title_list.index(title)) + 1
                    self.write_excel_detail(save_path, text_id, times, title, content_url)
                    print('已更新文章数据：' + title)
                else:   # 此时标题不在列表中
                    # 此时current_id处为空行
                    times = i[0]
                    title = i[1]
                    content_url = i[2]
                    self.write_excel_detail(save_path, current_id, times, title, content_url)
                    print('在序号为：' + str(current_id) + ' 处添加新文章：' + title)
                    current_id += 1
                    print('添加完后，序号变为：' + str(current_id))
        print('\n文章链接等信息已保存在：' + save_path)

    def read_excel(self, data_path):
        # 打开Excel文件
        workbook = openpyxl.load_workbook(data_path)
        # 获取活动工作表（通常是第一个工作表）
        sheet = workbook.active
        # 读取整行或整列数据
        # row_values = [cell.value for cell in sheet[1]]  # 第一行数据
        column_id = [sheet[f'A{i}'].value for i in range(2, sheet.max_row + 1)]  # A列数据
        column_values = [sheet[f'C{i}'].value for i in range(2, sheet.max_row + 1)]  # C列数据
        return column_id, column_values

    def creat_excel(self, save_path, pxlsx):
        if os.path.exists(save_path + pxlsx):   # 检测文件是否存在，若不存在则新创建
            print("列表文件存在")
            column_id, title_list = self.read_excel(save_path + pxlsx)
            if title_list:  # 表中有数据
                return {'exist_excel': 1}
            else:   # 表中无数据，重置表格内容
                print('注意！表中无数据，将重置表格内容')
                wb = openpyxl.Workbook()
                wb.save(save_path + pxlsx)
                array = ['序号', '发布时间', '文章名称', '文章链接']
                array_1 = ['A', 'B', 'C', 'D']
                for i in range(0, len(array)):
                    w = openpyxl.load_workbook(save_path + pxlsx)
                    sheet = w.active
                    sheet[array_1[i] + "1"] = array[i]
                    w.save(save_path + pxlsx)
                return {'exist_excel': 0}
        else:
            print('文件不存在')
            wb = openpyxl.Workbook()
            wb.save(save_path + pxlsx)
            array = ['序号', '发布时间', '文章名称', '文章链接']
            array_1 = ['A', 'B', 'C', 'D']
            for i in range(0, len(array)):
                w = openpyxl.load_workbook(save_path + pxlsx)
                sheet = w.active
                sheet[array_1[i] + "1"] = array[i]
                w.save(save_path + pxlsx)
            return {'exist_excel': 0}

    def write_excel_detail(self, path, orders, times, titles, links):
        wb = openpyxl.load_workbook(path)
        sheet = wb.active
        sheet['A' + str(orders + 1)] = orders
        sheet['B' + str(orders + 1)] = times
        sheet['C' + str(orders + 1)] = titles
        sheet['D' + str(orders + 1)] = links
        wb.save(path)
        order = orders + 1
        return order

    def creat_content(self,save_path, pxlsx):
        wb = openpyxl.Workbook()
        wb.save(save_path + pxlsx)
        array = ['序号', '时间', '文章名称', '阅读量', '点赞数', '转发数', '在看数', '文章链接', '评论', '评论点赞', '文章内容']
        array_1 = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K']
        for i in range(0, len(array)):
            w = openpyxl.load_workbook(save_path + pxlsx)
            sheet = w.active
            sheet[array_1[i] + "1"] = array[i]
            w.save(save_path + pxlsx)

    def write_content(self, path, orders, times, titles, read_nums, like_nums, share_nums, looking_nums, links,
                           comment, comment_likes, text_content):
        wb = openpyxl.load_workbook(path)
        sheet = wb.active
        sheet['A' + str(orders + 1)] = orders
        sheet['B' + str(orders + 1)] = times
        sheet['C' + str(orders + 1)] = titles
        sheet['D' + str(orders + 1)] = read_nums
        sheet['E' + str(orders + 1)] = like_nums
        sheet['F' + str(orders + 1)] = share_nums
        sheet['G' + str(orders + 1)] = looking_nums
        sheet['H' + str(orders + 1)] = links
        sheet['I' + str(orders + 1)] = comment
        sheet['J' + str(orders + 1)] = comment_likes
        sheet['K' + str(orders + 1)] = text_content
        wb.save(path)
        order = orders + 1
        return order



