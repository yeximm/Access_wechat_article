import requests
import json
import jsonpath
import random
import re
import openpyxl
from save_content import GetList


class SaveAllDetail(GetList):
    def __init__(self):
        super().__init__()

    def get_detail_new(self, link,title):
        r, appmsg_type, mid, sn, idx, ct, comment_id, version, req_id, detail_time, texts = self.get_Alltype(link)
        # url = ('https://mp.weixin.qq.com/mp/getappmsgext?f=json&mock=&fasttmplajax=1&f=json&uin='
        #        + self.uin + '&key=' + self.key + '&pass_ticket=' + self.pass_ticket
        #        + '&wxtoken=&devicetype=Windows%2B7%2Bx64&clientversion=6309092b'
        url = ('https://mp.weixin.qq.com/mp/getappmsgext?f=json&mock=&fasttmplajax=1&f=json'
               '&uin=' + self.uin + '&key=' + self.key + '&pass_ticket=' + self.pass_ticket + '&__biz=' + self.biz)
        data = {
            'r': r,
            '__biz': self.biz,
            'appmsg_type': '9',
            'mid': mid,
            'sn': sn,
            'idx': idx,
            'scene': '38',
            'title': title,
            'abtest_cookie': '',
            'devicetype': 'Windows 7 x64',
            'version': '63090b13',
            'is_need_ticket': '0',
            'is_need_ad': '0',
            'comment_id': comment_id,
            'is_need_reward': '0',
            'both_ad': '0',
            'reward_uin_count': '0',
            'send_time': '',
            'msg_daily_idx': '1',
            'is_original': '0',
            'is_only_read': '1',
            'req_id': req_id,
            'pass_ticket': self.pass_ticket,
            'is_temp_url': '0',
            'item_show_type': '0',
            'tmp_version': '1',
            'more_read_type': '0',
            'appmsg_like_type': '2',
            'related_video_sn': '',
            'related_video_num': '5',
            'vid': '',
            'is_pay_subscribe': '0',
            'pay_subscribe_uin_count': '0',
            'has_red_packet_cover': '0',
            'album_id': '1296223588617486300',
            'album_video_num': '5',
            'cur_album_id': 'undefined',
            'is_public_related_video': 'NaN',
            'encode_info_by_base64': 'undefined',
            'exptype': '',
            'export_key_extinfo': '',
            'business_type': '0',
        }
        # print(url,data)
        res = requests.post(url=url, data=data, headers=self.headers, cookies=self.cookies, verify=False)
        # print(res.text)
        read_num = self.get_json(res.text, "read_num")
        like_num = self.get_json(res.text, "old_like_num")
        share_num = self.get_json(res.text, "share_num")
        show_read = self.get_json(res.text, "show_read")
        # print(f'read_num:{read_num},like_num:{like_num}', 'get_detail_new')
        comments,like_nums=self.get_comment(comment_id)
        if read_num==[] or read_num=='':
            return '','','',''
        else:
            return read_num[0], like_num[0], share_num[0], show_read[0],comments,like_nums,detail_time,title,texts

    def get_json(self,text,json_word):
        content = json.loads(text)
        value = jsonpath.jsonpath(content, "$.."+json_word)
        return value

    def get_comment(self, comment_id):
        url = ('https://mp.weixin.qq.com/mp/appmsg_comment?action=getcomment&__biz=' + self.biz +
                '&appmsgid=2247491372&idx=1&comment_id=' + comment_id + '&offset=0&limit=100&uin='
               + self.uin + '&key=' + self.key + '&pass_ticket=' + self.pass_ticket
               + '&wxtoken=&devicetype=Windows+10&clientversion=62060833&appmsg_token=')
        response = requests.get(url, headers=self.headers, cookies=self.cookies, verify=False)
        json_content = json.loads(response.text)
        content = jsonpath.jsonpath(json_content, '$..content')
        like_nums = jsonpath.jsonpath(json_content, '$..like_num')
        return content, like_nums

    def get_Alltype(self, link):
        # url = link
        # res = requests.get(url=link, headers=self.header, cookies=self.cookie, verify=False)
        contents = self.get_content(link)   # 继承SaveContent()类的方法
        if contents['content_flag'] == 0:
            # 文章获取失败
            return {'content_flag': 0}
        else:
            detail_time = self.get_time(contents['content'])  # 获取文章发布时间信息
            texts = self.get_texts(contents['content'])  # 列表形式的文章内容
            tittle = self.get_title(contents['content'])
            # print(detail_time, tittle, texts)
            r = ''
            for rand in range(0, 16):
                r = '' + str(random.randint(0, 9))
            a = ('http://mp.weixin.qq.com/s?'
             '__biz=MzkxMDMxODEwOA==&'
             'amp;mid=2247491478&'
             'amp;idx=1&'
             'amp;sn=cd91230a7d6db0783e30b9d62fc65d8b&'
             'amp;chksm=c01d1272e5a57923333dbcf83e0ab7ee799cdb3d2074d0d3e1889fbdd85b2b1e1ecefca7f876&'
             'amp;scene=27')
            r = '0.' + r
            appmsg_type = "9"
            mid = str(link).split('mid=')[1].split('&')[0]
            sn = str(link).split('sn=')[1].split('&')[0]
            idx = str(link).split('idx=')[1].split('&')[0]
            ct = ''
            # titles = t
            # title = urllib.parse.quote(str(titles).encode('utf-8'))
            # print(res.text)
            # comment_id = res.text.split('var comment_id = "')[1].split('"')[0].replace("'", "").replace('"', '')
            comment_id = re.search('var comment_id.*"(.*?)"',contents['content'])
            if comment_id:
                comment_id = re.search('var comment_id.*"(.*?)"', contents['content']).group(1)
            else:
                # comment_id = re.search('d.comment_id =(.*?);', res.text).group(1)
                print('没有匹配到comment_id，文章标题为：' + tittle)
                comment_id = ''
            version = contents['content'].split('_g.clientversion = "')[1].split('"')[0]
            if 'var req_id = ' in contents['content']:
                req_id = contents['content'].split('var req_id = ')[1].split(';')[0].replace("'", "").replace('"', '')
            else:
                print('没有匹配到req_id，文章标题为：' + tittle)
                req_id = ''
            # return r, biz, appmsg_type, mid, sn, idx, ct, title, comment_id, version, req_id
            return r, appmsg_type, mid, sn, idx, ct, comment_id, version, req_id, detail_time, texts

    def creat_excel_detail(self,save_path, pxlsx):
        wb = openpyxl.Workbook()
        wb.save(save_path + pxlsx)
        array = ['序号', '时间', '文章名称', '阅读量', '点赞数', '转发数', '在看数', '文章链接', '评论', '评论点赞', '文章内容']
        array_1 = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K']
        for i in range(0, len(array)):
            w = openpyxl.load_workbook(save_path + pxlsx)
            sheet = w.active
            sheet[array_1[i] + "1"] = array[i]
            w.save(save_path + pxlsx)

    def write_excel_detail(self, path, orders, times, titles, read_nums, like_nums, share_nums, looking_nums, links,
                           comment, comment_likes, text_content):
        wb = openpyxl.load_workbook(path)
        sheet = wb.active
        sheet['A' + str(orders + 1)] = orders
        sheet['B' + str(orders + 1)] = times
        sheet['C' + str(orders + 1)] = titles
        sheet['D' + str(orders + 1)] = read_nums
        sheet['E' + str(orders + 1)] = like_nums
        sheet['F' + str(orders + 1)] = share_nums
        sheet['G' + str(orders + 1)] = looking_nums
        sheet['H' + str(orders + 1)] = links
        sheet['I' + str(orders + 1)] = comment
        sheet['J' + str(orders + 1)] = comment_likes
        sheet['K' + str(orders + 1)] = text_content
        wb.save(path)
        order = orders + 1
        return order
